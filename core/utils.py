# -*- coding: utf-8 -*-

import os
import csv

from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference

def load_data_from_excel(path, sheetname, lo, hi):
    """ Загрузка коэффициентов усиления из excel"""
    wb = load_workbook(path, data_only=True)
    sheet = wb[sheetname]
    result = []
    for cellobj in sheet[lo:hi]:
        for cell in cellobj:
            result.append(cell.value)
    return result


def load_potentiometer_gain(path):
    res = []
    with open(path, 'r') as f:
        for line in f:
            line = line.strip('\n ,')
            res.extend([float(x) for x in line.split(',')])
    return res


def to_csv(filename, data, mode='w', newline=''):
    with open(filename, mode, newline=newline) as f:
        writer = csv.writer(f, delimiter=',')
        writer.writerows(data)


def to_excel(path, data, tolerance, ratio_m, headers=None, chart=True):
    wb = Workbook()
    ws = wb.active
    
    if headers:
        ws.append(headers)

    for row in data:
        ws.append(row)
        
    max_row = len(data)

    chart  = LineChart()
    chart.title = f"Gain (|S1-S2| = {tolerance}, M = {ratio_m})"    
    chart.x_axis.title = "Step"
    chart.y_axis.title = "Gain"
    data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=max_row)
    data2 = Reference(ws, min_col=3, min_row=1, max_col=3, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.add_data(data2, titles_from_data=True)
    ws.add_chart(chart, "L1")

    chart  = LineChart()
    chart.title = f"КУ - K1 x K2"    
    chart.x_axis.title = "Step"
    chart.y_axis.title = "Deviation"
    data = Reference(ws, min_col=6, min_row=1, max_col=6, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    ws.add_chart(chart, "L15")

    try:
        wb.save(path)
    except PermissionError:
        raise PermissionError

def make_directory(path):
    if not os.path.exists(path):
        os.mkdir(path)
    

if __name__ == "__main__":
    to_excel("", [])