# -*- coding: utf-8 -*-

import os
from datetime import datetime

from openpyxl import load_workbook

from PySide2.QtGui import QPixmap
from PySide2.QtWidgets import *
from PySide2 import QtCore
from PySide2 import QtCharts

from ui.charts import LineChart, DeviationChart
from ui.preferencesdialog import Ui_Dialog

from core import utils


class OpenDialog(QDialog):
    def __init__(self, parent, stg=None) -> None:
        super(OpenDialog, self).__init__(parent)

        self.ui = Ui_Dialog()
        self.ui.setupUi(self)

    # Connect signal/slot
        self.ui.browseButton.clicked.connect(self.on_browse)

    def on_browse(self):
        filename, _ = QFileDialog.getOpenFileName(
            self, "Open file", "./data", "Data Files (*.xlsx)"
        )
        if not filename:
            return

        self.ui.filenameBox.setText(filename)

        wb = load_workbook(filename, data_only=True)
        self.sheetnames = wb.sheetnames
        self.ui.gainSheetBox.addItems(self.sheetnames)
        self.ui.ratioSheetBox.addItems(self.sheetnames)

    def read_settings(self):
        return {
            "filename": self.ui.filenameBox.text(),
            "gain": {
                "sheet": self.ui.gainSheetBox.currentText(),
                "column": self.ui.gainColumnBox.text(),
                "min_row": self.ui.gainRangeMinBox.text(),
                "max_row": self.ui.gainRangeMaxBox.text()
            },
            "ratio": {
                "sheet": self.ui.ratioSheetBox.currentText(),
                "column": self.ui.ratioColumnBox.text(),
                "min_row": self.ui.ratioRangeMinBox.text(),
                "max_row": self.ui.ratioRangeMaxBox.text()
            }
        }



class ChartDialog(QDialog):
    def __init__(self, model=None) -> None:
        super(ChartDialog, self).__init__()

        self.model = model

        self.setModal(True)
        self.setWindowFlag(QtCore.Qt.WindowMaximizeButtonHint)
        self.setSizeGripEnabled(True)
        self.resize(800, 800)

        self.setupUI()

        self.saveButton.clicked.connect(self.on_save)


    def setupUI(self):
        self._primary_chart = LineChart(self.model)
        self._primary_chart.setWindowTitle("gains")

        self._extra_chart = DeviationChart(self.model)
        self._extra_chart.setWindowTitle("deviation")
        self._extra_chart.setMaximumHeight(self.height() / 2 - 1)

        layout = QVBoxLayout(self)
        hbox = QHBoxLayout()

        self.saveButton = QPushButton(self.tr("Save"))
        hbox.addWidget(self.saveButton)

        self.clearButton = QPushButton(self.tr("Clear"))
        hbox.addWidget(self.clearButton)

        spacer = QSpacerItem(40, 20, QSizePolicy.Expanding, QSizePolicy.Minimum)
        hbox.addSpacerItem(spacer)

        layout.addLayout(hbox)
        layout.addWidget(self._primary_chart)
        layout.addWidget(self._extra_chart)

    def on_save(self):    
        output = "./output"
        utils.make_directory(output)
        for chart in (w for w in self.children() if isinstance(w, QtCharts.QChartView)): 
            pixmap = QPixmap(chart.size())
            chart.render(pixmap)
            filename = datetime.now().strftime("%Y%m%d")  + "_" + chart.windowTitle()
            path = os.path.join(output, filename + ".png")
            print(path)
            pixmap.save(path, "PNG")


if __name__ == "__main__":
    import sys
    app = QApplication([])
    mw = OpenDialog()
    mw.show()
    sys.exit(app.exec_())


        